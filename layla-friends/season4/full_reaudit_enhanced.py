#!/usr/bin/env python3
"""
Enhanced full reaudit of Season 4 scripts with Excel and Markdown output.
Verifies tier progression and generates comprehensive reports.
"""

import os
import re
import sys
from collections import defaultdict
import json
import string

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Skipping Excel output.")

def count_syllables(word):
    """Approximate syllable count for a word."""
    word = word.lower().strip(string.punctuation)
    if len(word) <= 3:
        return 1

    vowels = 'aeiouy'
    syllable_count = 0
    previous_was_vowel = False

    for char in word:
        is_vowel = char in vowels
        if is_vowel and not previous_was_vowel:
            syllable_count += 1
        previous_was_vowel = is_vowel

    # Adjust for silent e
    if word.endswith('e'):
        syllable_count -= 1

    # Ensure at least 1 syllable
    if syllable_count == 0:
        syllable_count = 1

    return syllable_count

def count_sentences(text):
    """Count sentences in text."""
    # Remove markdown headers and metadata
    text = re.sub(r'^#+\s+.*$', '', text, flags=re.MULTILINE)
    text = re.sub(r'\*\*.*?\*\*', '', text)

    # Count sentence-ending punctuation
    sentences = re.findall(r'[.!?]+', text)
    return max(len(sentences), 1)  # At least 1 sentence

def calculate_flesch_kincaid(content):
    """Calculate Flesch-Kincaid grade level from script content."""
    # Extract script content (scenes only, ignore metadata)
    # Try multiple scene formats:
    # 1. "### Scene \d+" (standard markdown header)
    # 2. "**SCENE \d+**" (all caps bold)
    # 3. "**Scene \d+**" (title case bold)

    scene_pattern = r'### Scene \d+.*?\n(.*?)(?=### Scene \d+|---|\Z)'
    scenes = re.findall(scene_pattern, content, re.DOTALL)

    if not scenes:
        # Try bold SCENE format
        scene_pattern = r'\*\*SCENE \d+\*\*\n(.*?)(?=\*\*SCENE \d+\*\*|---|\Z)'
        scenes = re.findall(scene_pattern, content, re.DOTALL)

    if not scenes:
        # Try bold Scene format
        scene_pattern = r'\*\*Scene \d+\*\*\n(.*?)(?=\*\*Scene \d+\*\*|---|\Z)'
        scenes = re.findall(scene_pattern, content, re.DOTALL)

    if not scenes:
        return None

    script_text = ' '.join(scenes)

    # Count words
    words = re.findall(r'\b[a-zA-Z]+\b', script_text)
    total_words = len(words)

    if total_words == 0:
        return None

    # Count syllables
    total_syllables = sum(count_syllables(word) for word in words)

    # Count sentences
    total_sentences = count_sentences(script_text)

    # Calculate Flesch-Kincaid grade level
    # F-K = 0.39 × (total words / total sentences) + 11.8 × (total syllables / total words) - 15.59
    fk_grade = 0.39 * (total_words / total_sentences) + 11.8 * (total_syllables / total_words) - 15.59

    return round(fk_grade, 2)

def extract_word_count(filepath):
    """Extract target and actual word counts from script."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    # Extract target word count
    target_match = re.search(r'\*\*Target Word Count:\*\* (\d+)', content)
    target_wc = int(target_match.group(1)) if target_match else None

    # Extract actual word count - try multiple formats
    actual_match = re.search(r'\*\*Total Word Count:? ([\d,]+) words\*\*', content)
    if actual_match:
        actual_wc = int(actual_match.group(1).replace(',', ''))
    else:
        # Try alternate format: "**WORD COUNT: 718 words**" or "**Word Count: 657 words**"
        actual_match = re.search(r'\*\*(?:WORD COUNT|Word Count):? ([\d,]+) words?\*\*', content, re.IGNORECASE)
        if actual_match:
            actual_wc = int(actual_match.group(1).replace(',', ''))
        else:
            actual_wc = None

    return target_wc, actual_wc

def extract_msl(filepath):
    """Extract mean sentence length from script."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    # Try both "Estimated MSL" and "Actual MSL" formats
    msl_match = re.search(r'\*\*(Estimated|Actual) MSL:\*\* ~?([\d.]+)', content)
    if msl_match:
        return float(msl_match.group(2))
    return None

def check_tier_progression(scripts_dir):
    """Check that each tier progresses properly within each chapter."""
    issues = []
    chapters = defaultdict(dict)

    # Get all script files
    script_files = []
    for filename in os.listdir(scripts_dir):
        if filename.startswith('S4-CH') and filename.endswith('.md'):
            script_files.append(filename)

    for filename in script_files:
        # Parse chapter and tier
        match = re.match(r'S4-CH(\d+)_TIER(\d)', filename)
        if not match:
            continue

        chapter = int(match.group(1))
        tier = int(match.group(2))

        filepath = os.path.join(scripts_dir, filename)
        target_wc, actual_wc = extract_word_count(filepath)
        msl = extract_msl(filepath)

        # Read content for F-K calculation
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        fk_grade = calculate_flesch_kincaid(content)

        chapters[chapter][tier] = {
            'filename': filename,
            'target_wc': target_wc,
            'actual_wc': actual_wc,
            'msl': msl,
            'fk_grade': fk_grade
        }

    # Check each chapter for proper progression
    for chapter in sorted(chapters.keys()):
        tiers = chapters[chapter]

        for tier in [1, 2, 3, 4]:
            if tier not in tiers:
                issues.append(f"CH{chapter:02d}: Missing TIER{tier}")
                continue

        # Check word count progression
        for tier in [1, 2, 3]:
            next_tier = tier + 1
            if tier in tiers and next_tier in tiers:
                curr_wc = tiers[tier]['actual_wc']
                next_wc = tiers[next_tier]['actual_wc']

                if curr_wc and next_wc and curr_wc >= next_wc:
                    issues.append(
                        f"CH{chapter:02d}: TIER{tier} word count ({curr_wc}) >= TIER{next_tier} ({next_wc})"
                    )

        # Check MSL progression
        for tier in [1, 2, 3]:
            next_tier = tier + 1
            if tier in tiers and next_tier in tiers:
                curr_msl = tiers[tier]['msl']
                next_msl = tiers[next_tier]['msl']

                if curr_msl and next_msl and curr_msl >= next_msl:
                    issues.append(
                        f"CH{chapter:02d}: TIER{tier} MSL ({curr_msl}) >= TIER{next_tier} ({next_msl})"
                    )

    return issues, chapters

def calculate_tier_averages(chapters):
    """Calculate average word counts, MSL, and F-K by tier across all chapters."""
    tier_stats = {1: [], 2: [], 3: [], 4: []}
    tier_msl = {1: [], 2: [], 3: [], 4: []}
    tier_fk = {1: [], 2: [], 3: [], 4: []}

    for chapter in chapters.values():
        for tier, data in chapter.items():
            if data['actual_wc']:
                tier_stats[tier].append(data['actual_wc'])
            if data['msl']:
                tier_msl[tier].append(data['msl'])
            if data['fk_grade']:
                tier_fk[tier].append(data['fk_grade'])

    averages = {}
    for tier in [1, 2, 3, 4]:
        wc_avg = sum(tier_stats[tier]) / len(tier_stats[tier]) if tier_stats[tier] else 0
        msl_avg = sum(tier_msl[tier]) / len(tier_msl[tier]) if tier_msl[tier] else 0
        fk_avg = sum(tier_fk[tier]) / len(tier_fk[tier]) if tier_fk[tier] else 0
        averages[tier] = {
            'word_count_avg': round(wc_avg, 1),
            'msl_avg': round(msl_avg, 2),
            'fk_avg': round(fk_avg, 2),
            'num_scripts': len(tier_stats[tier])
        }

    return averages

def generate_markdown_report(chapters, issues, averages, output_path):
    """Generate markdown summary report."""
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("# Season 4 Full Reaudit Summary\n\n")
        f.write("## Tier Progression Status\n\n")

        if issues:
            f.write(f"⚠️ **{len(issues)} ISSUES FOUND**\n\n")
            for issue in issues:
                f.write(f"- ❌ {issue}\n")
            f.write("\n")
        else:
            f.write("✅ **NO ISSUES FOUND**\n\n")
            f.write("- All chapters have proper tier progression\n")
            f.write("- Word counts increase across tiers\n")
            f.write("- MSL increases across tiers\n\n")

        f.write("## Tier Averages Across All Chapters\n\n")
        f.write("| Tier | Avg Word Count | Avg MSL | Avg F-K Grade | Scripts |\n")
        f.write("|------|----------------|---------|---------------|----------|\n")
        for tier in [1, 2, 3, 4]:
            avg = averages[tier]
            f.write(f"| T{tier}  | {avg['word_count_avg']:>7.1f} | {avg['msl_avg']:>7.2f} | {avg['fk_avg']:>13.2f} | {avg['num_scripts']:>8} |\n")

        f.write("\n## Chapter-by-Chapter Breakdown\n\n")
        for chapter in sorted(chapters.keys()):
            f.write(f"### Chapter {chapter:02d}\n\n")
            f.write("| Tier | Word Count | MSL | F-K Grade | Target WC | File |\n")
            f.write("|------|-----------|-----|-----------|-----------|------|\n")

            tiers = chapters[chapter]
            for tier in [1, 2, 3, 4]:
                if tier in tiers:
                    data = tiers[tier]
                    wc = data['actual_wc'] or 'N/A'
                    msl = data['msl'] or 'N/A'
                    fk = data['fk_grade'] or 'N/A'
                    target = data['target_wc'] or 'N/A'
                    f.write(f"| T{tier} | {wc:>9} | {msl:>7} | {fk:>9} | {target:>9} | {data['filename']} |\n")
                else:
                    f.write(f"| T{tier} | MISSING | MISSING | MISSING | MISSING | - |\n")
            f.write("\n")

def generate_excel_report(chapters, issues, averages, output_path):
    """Generate Excel report with detailed statistics."""
    if not HAS_OPENPYXL:
        print("Skipping Excel output (openpyxl not installed)")
        return

    wb = openpyxl.Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Headers
    ws_summary['A1'] = "Season 4 Full Reaudit"
    ws_summary['A1'].font = Font(size=16, bold=True)

    ws_summary['A3'] = "Status:"
    if issues:
        ws_summary['B3'] = f"{len(issues)} ISSUES FOUND"
        ws_summary['B3'].fill = PatternFill(start_color="FFCCCC", fill_type="solid")
    else:
        ws_summary['B3'] = "NO ISSUES - ALL PASSING"
        ws_summary['B3'].fill = PatternFill(start_color="CCFFCC", fill_type="solid")

    # Tier averages table
    ws_summary['A5'] = "Tier Averages"
    ws_summary['A5'].font = Font(bold=True)

    ws_summary['A6'] = "Tier"
    ws_summary['B6'] = "Avg Word Count"
    ws_summary['C6'] = "Avg MSL"
    ws_summary['D6'] = "Avg F-K Grade"
    ws_summary['E6'] = "# Scripts"

    for i, header in enumerate(['A6', 'B6', 'C6', 'D6', 'E6']):
        ws_summary[header].font = Font(bold=True)
        ws_summary[header].fill = PatternFill(start_color="DDDDDD", fill_type="solid")

    for idx, tier in enumerate([1, 2, 3, 4], start=7):
        avg = averages[tier]
        ws_summary[f'A{idx}'] = f"T{tier}"
        ws_summary[f'B{idx}'] = avg['word_count_avg']
        ws_summary[f'C{idx}'] = avg['msl_avg']
        ws_summary[f'D{idx}'] = avg['fk_avg']
        ws_summary[f'E{idx}'] = avg['num_scripts']

    # Issues sheet if any
    if issues:
        ws_issues = wb.create_sheet("Issues")
        ws_issues['A1'] = "Issue"
        ws_issues['A1'].font = Font(bold=True)
        for idx, issue in enumerate(issues, start=2):
            ws_issues[f'A{idx}'] = issue

    # Detailed data sheet
    ws_detail = wb.create_sheet("All Scripts")
    headers = ['Chapter', 'Tier', 'Word Count', 'MSL', 'F-K Grade', 'Target WC', 'Filename']
    for idx, header in enumerate(headers, start=1):
        cell = ws_detail.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")

    row = 2
    for chapter in sorted(chapters.keys()):
        tiers = chapters[chapter]
        for tier in [1, 2, 3, 4]:
            if tier in tiers:
                data = tiers[tier]
                ws_detail.cell(row=row, column=1, value=chapter)
                ws_detail.cell(row=row, column=2, value=tier)
                ws_detail.cell(row=row, column=3, value=data['actual_wc'])
                ws_detail.cell(row=row, column=4, value=data['msl'])
                ws_detail.cell(row=row, column=5, value=data['fk_grade'])
                ws_detail.cell(row=row, column=6, value=data['target_wc'])
                ws_detail.cell(row=row, column=7, value=data['filename'])
                row += 1

    wb.save(output_path)

def main():
    print("Running Season 4 Full Reaudit (Enhanced)...")

    # Determine scripts directory
    if len(sys.argv) > 1:
        # Path provided as argument
        scripts_dir = os.path.dirname(sys.argv[1]) if os.path.dirname(sys.argv[1]) else 'season4/scripts'
    else:
        scripts_dir = 'season4/scripts'

    if not os.path.exists(scripts_dir):
        print(f"Error: Scripts directory not found: {scripts_dir}")
        return 1

    print(f"Checking tier progression in: {scripts_dir}")

    issues, chapters = check_tier_progression(scripts_dir)
    averages = calculate_tier_averages(chapters)

    # Print console output
    print("\n" + "="*90)
    print("TIER AVERAGES ACROSS ALL CHAPTERS")
    print("="*90)
    print(f"{'Tier':<6} {'Avg Word Count':>15} {'Avg MSL':>10} {'Avg F-K Grade':>15} {'Scripts':>10}")
    print("-"*90)
    for tier in [1, 2, 3, 4]:
        avg = averages[tier]
        print(f"T{tier}    {avg['word_count_avg']:>15.1f} {avg['msl_avg']:>10.2f} {avg['fk_avg']:>15.2f} {avg['num_scripts']:>10}")

    print("\n" + "="*80)
    print("TIER PROGRESSION ISSUES")
    print("="*80)

    if issues:
        print(f"\n⚠️  Found {len(issues)} progression issue(s):\n")
        for issue in issues:
            print(f"  ❌ {issue}")
        print("\nFIX REQUIRED: Progression inversions detected!")
        status = 1
    else:
        print("\n✅ No progression issues found!")
        print("✅ All chapters have proper tier progression")
        print("✅ Word counts increase across tiers")
        print("✅ MSL increases across tiers")
        status = 0

    # Generate reports
    output_dir = 'season4'
    md_path = os.path.join(output_dir, 'REAUDIT_SUMMARY.md')
    xlsx_path = os.path.join(output_dir, 'FULL_REAUDIT.xlsx')

    print(f"\nGenerating reports...")
    generate_markdown_report(chapters, issues, averages, md_path)
    print(f"✅ Markdown report: {md_path}")

    generate_excel_report(chapters, issues, averages, xlsx_path)
    if HAS_OPENPYXL:
        print(f"✅ Excel report: {xlsx_path}")

    return status

if __name__ == '__main__':
    exit(main())
