#!/usr/bin/env python3
"""
Season 3 Full Audit with Flesch-Kincaid Grade Level
Checks tier progression and reading level metrics
"""

import re
import os
import string
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def count_syllables(word):
    """Approximate syllable count using vowel patterns."""
    word = word.lower().strip(string.punctuation)
    if len(word) <= 3:
        return 1
    vowels = 'aeiouy'
    syllable_count = 0
    previous_was_vowel = False
    for char in word:
        is_vowel = char in vowels
        if is_vowel and not previous_was_vowel:
            syllable_count += 1
        previous_was_vowel = is_vowel
    if word.endswith('e'):
        syllable_count -= 1
    if syllable_count == 0:
        syllable_count = 1
    return syllable_count

def count_sentences(text):
    """Count sentences in text."""
    sentences = re.split(r'[.!?]+', text)
    return len([s for s in sentences if s.strip()])

def calculate_flesch_kincaid(content):
    """Calculate Flesch-Kincaid grade level."""
    lines = content.split('\n')
    scene_content = []
    in_scene = False

    for line in lines:
        if line.startswith('**Scene') or line.startswith('**SCENE') or line.startswith('### Scene'):
            in_scene = True
            continue
        if line.startswith('---') or line.startswith('**Word Count') or line.startswith('**WORD COUNT') or line.startswith('**Total Word Count'):
            in_scene = False
        if in_scene and line.strip():
            scene_content.append(line)

    text = ' '.join(scene_content)
    words = text.split()
    total_words = len(words)

    if total_words == 0:
        return 0.0

    total_sentences = count_sentences(text)
    if total_sentences == 0:
        total_sentences = 1

    total_syllables = sum(count_syllables(word) for word in words)

    fk_grade = 0.39 * (total_words / total_sentences) + 11.8 * (total_syllables / total_words) - 15.59
    return round(fk_grade, 2)

def extract_word_count(content):
    """Extract word count from script metadata."""
    # Try multiple patterns - allow for commas in numbers
    patterns = [
        r'\*\*Total Word Count[:\s]+([\d,]+)',
        r'\*\*Word Count[:\s]+([\d,]+)',
        r'WORD COUNT[:\s]+([\d,]+)',
        r'Word Count[:\s]+([\d,]+)',
    ]

    for pattern in patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            # Remove commas and convert to int
            return int(match.group(1).replace(',', ''))
    return None

def extract_msl(content):
    """Extract MSL from script metadata."""
    # Try multiple patterns
    patterns = [
        r'\*\*Estimated MSL[:\s]+~?(\d+\.?\d*)',
        r'\*\*Actual MSL[:\s]+~?(\d+\.?\d*)',
        r'MSL[:\s]+~?(\d+\.?\d*)',
    ]

    for pattern in patterns:
        match = re.search(pattern, content, re.IGNORECASE)
        if match:
            return float(match.group(1))
    return None

def check_tier_progression(scripts_dir):
    """Check tier progression for all chapters."""
    print(f"Checking tier progression in: {scripts_dir}\n")

    all_results = []
    issues = []

    for ch in range(1, 13):
        ch_num = f"{ch:02d}"

        # Get file paths for all tiers
        tier_files = {
            1: scripts_dir / f"S3-CH{ch_num}_TIER1.md",
            2: scripts_dir / f"S3-CH{ch_num}_TIER2.md",
            3: scripts_dir / f"S3-CH{ch_num}_TIER3.md",
            4: scripts_dir / f"S3-CH{ch_num}_TIER4.md",
        }

        tier_data = {}

        for tier, filepath in tier_files.items():
            if not filepath.exists():
                continue

            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()

            word_count = extract_word_count(content)
            msl = extract_msl(content)
            fk_grade = calculate_flesch_kincaid(content)

            tier_data[tier] = {
                'word_count': word_count,
                'msl': msl,
                'fk_grade': fk_grade,
                'file': filepath.name
            }

            all_results.append({
                'chapter': ch,
                'tier': tier,
                'file': filepath.name,
                'word_count': word_count,
                'msl': msl,
                'fk_grade': fk_grade
            })

        # Check progression
        for tier in range(1, 4):
            if tier not in tier_data or tier + 1 not in tier_data:
                continue

            curr = tier_data[tier]
            next_tier = tier_data[tier + 1]

            # Check word count progression
            if curr['word_count'] and next_tier['word_count']:
                if curr['word_count'] >= next_tier['word_count']:
                    issues.append(f"CH{ch_num}: T{tier} word count ({curr['word_count']}) >= T{tier+1} ({next_tier['word_count']})")

            # Check MSL progression
            if curr['msl'] and next_tier['msl']:
                if curr['msl'] >= next_tier['msl']:
                    issues.append(f"CH{ch_num}: T{tier} MSL ({curr['msl']}) >= T{tier+1} ({next_tier['msl']})")

    return all_results, issues

def calculate_tier_averages(results):
    """Calculate average metrics for each tier."""
    tier_averages = {}

    for tier in [1, 2, 3, 4]:
        tier_results = [r for r in results if r['tier'] == tier]
        if not tier_results:
            continue

        avg_wc = sum(r['word_count'] for r in tier_results if r['word_count']) / len([r for r in tier_results if r['word_count']])
        avg_msl = sum(r['msl'] for r in tier_results if r['msl']) / len([r for r in tier_results if r['msl']])
        avg_fk = sum(r['fk_grade'] for r in tier_results if r['fk_grade']) / len([r for r in tier_results if r['fk_grade']])

        tier_averages[tier] = {
            'avg_word_count': round(avg_wc, 1),
            'avg_msl': round(avg_msl, 2),
            'avg_fk_grade': round(avg_fk, 2),
            'count': len(tier_results)
        }

    return tier_averages

def generate_reports(results, tier_averages, issues, output_dir):
    """Generate markdown and Excel reports."""

    # Markdown report
    md_path = output_dir / "S3_FULL_AUDIT.md"
    with open(md_path, 'w') as f:
        f.write("# Season 3 Full Audit Summary\n\n")

        # Tier averages
        f.write("## Tier Averages Across All Chapters\n\n")
        f.write("| Tier | Avg Word Count | Avg MSL | Avg F-K Grade | Scripts |\n")
        f.write("|------|----------------|---------|---------------|----------|\n")
        for tier in [1, 2, 3, 4]:
            if tier in tier_averages:
                ta = tier_averages[tier]
                f.write(f"| T{tier}  | {ta['avg_word_count']:7.1f} | {ta['avg_msl']:7.2f} | {ta['avg_fk_grade']:13.2f} | {ta['count']:8} |\n")

        # Chapter breakdown
        f.write("\n## Chapter-by-Chapter Breakdown\n\n")
        for ch in range(1, 13):
            ch_results = [r for r in results if r['chapter'] == ch]
            if not ch_results:
                continue

            f.write(f"\n### Chapter {ch:02d}\n\n")
            f.write("| Tier | Word Count | MSL | F-K Grade | File |\n")
            f.write("|------|-----------|-----|-----------|------|\n")

            for tier in [1, 2, 3, 4]:
                tier_result = next((r for r in ch_results if r['tier'] == tier), None)
                if tier_result:
                    wc = tier_result['word_count'] if tier_result['word_count'] else 'N/A'
                    msl = f"{tier_result['msl']:.1f}" if tier_result['msl'] else 'N/A'
                    fk = f"{tier_result['fk_grade']:.2f}" if tier_result['fk_grade'] else 'N/A'
                    f.write(f"| T{tier} | {wc:9} | {msl:7} | {fk:9} | {tier_result['file']} |\n")

        # Issues
        f.write("\n## Tier Progression Issues\n\n")
        if issues:
            for issue in issues:
                f.write(f"- ⚠️  {issue}\n")
        else:
            f.write("✅ No progression issues found!\n")

    # Excel report
    xlsx_path = output_dir / "S3_FULL_AUDIT.xlsx"
    wb = openpyxl.Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Headers
    headers = ["Tier", "Avg Word Count", "Avg MSL", "Avg F-K Grade", "Scripts"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Data
    for row, tier in enumerate([1, 2, 3, 4], 2):
        if tier in tier_averages:
            ta = tier_averages[tier]
            ws_summary.cell(row, 1, f"T{tier}")
            ws_summary.cell(row, 2, ta['avg_word_count'])
            ws_summary.cell(row, 3, ta['avg_msl'])
            ws_summary.cell(row, 4, ta['avg_fk_grade'])
            ws_summary.cell(row, 5, ta['count'])

    # All scripts sheet
    ws_all = wb.create_sheet("All Scripts")
    headers = ["Chapter", "Tier", "File", "Word Count", "MSL", "F-K Grade"]
    for col, header in enumerate(headers, 1):
        cell = ws_all.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for row, result in enumerate(sorted(results, key=lambda x: (x['chapter'], x['tier'])), 2):
        ws_all.cell(row, 1, f"CH{result['chapter']:02d}")
        ws_all.cell(row, 2, f"T{result['tier']}")
        ws_all.cell(row, 3, result['file'])
        ws_all.cell(row, 4, result['word_count'] if result['word_count'] else 'N/A')
        ws_all.cell(row, 5, result['msl'] if result['msl'] else 'N/A')
        ws_all.cell(row, 6, result['fk_grade'] if result['fk_grade'] else 'N/A')

    # Issues sheet (if any)
    if issues:
        ws_issues = wb.create_sheet("Issues")
        ws_issues.cell(1, 1, "Issue Description").font = Font(bold=True)
        for row, issue in enumerate(issues, 2):
            ws_issues.cell(row, 1, issue)

    wb.save(xlsx_path)

    return md_path, xlsx_path

def main():
    print("Running Season 3 Full Audit (Enhanced)...")

    scripts_dir = Path("season3/scripts")
    if not scripts_dir.exists():
        print(f"Error: Scripts directory not found: {scripts_dir}")
        return 1

    results, issues = check_tier_progression(scripts_dir)

    tier_averages = calculate_tier_averages(results)

    # Print tier averages
    print("=" * 90)
    print("TIER AVERAGES ACROSS ALL CHAPTERS")
    print("=" * 90)
    print(f"{'Tier':<8}{'Avg Word Count':>16}{'Avg MSL':>10}{'Avg F-K Grade':>16}{'Scripts':>12}")
    print("-" * 90)
    for tier in [1, 2, 3, 4]:
        if tier in tier_averages:
            ta = tier_averages[tier]
            print(f"T{tier:<7}{ta['avg_word_count']:>16.1f}{ta['avg_msl']:>10.2f}{ta['avg_fk_grade']:>16.2f}{ta['count']:>12}")

    # Print issues
    print("\n" + "=" * 80)
    print("TIER PROGRESSION ISSUES")
    print("=" * 80)
    print()
    if issues:
        for issue in issues:
            print(f"⚠️  {issue}")
    else:
        print("✅ No progression issues found!")
        print("✅ All chapters have proper tier progression")
        print("✅ Word counts increase across tiers")
        print("✅ MSL increases across tiers")

    # Generate reports
    print("\nGenerating reports...")
    md_path, xlsx_path = generate_reports(results, tier_averages, issues, Path("season3"))
    print(f"✅ Markdown report: {md_path}")
    print(f"✅ Excel report: {xlsx_path}")

    return 0

if __name__ == '__main__':
    exit(main())
